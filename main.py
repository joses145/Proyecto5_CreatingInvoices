import pandas as pd
import glob
from fpdf import FPDF
from pathlib import Path
import openpyxl

filepaths = glob.glob("invoices/*.xlsx")

for filepath in filepaths:
    df = pd.read_excel(filepath, sheet_name="Sheet 1")

    pdf = FPDF(orientation="P", unit="mm", format="A4")
    pdf.add_page()

    filename = Path(filepath).stem
    invoice_number, invoice_date = filename.split("-")

    pdf.set_font(family="Times", size=18, style="B")
    pdf.cell(w=50, h=8, txt=f"Invoice nr. {invoice_number}")

    pdf.set_font(family="Times", size=18, style="B")
    pdf.cell(w=50, h=8, txt=f"Date: {invoice_date}", ln=1)

    columns = list(df.columns)

    pdf.set_font(family="Arial", size=10, style="B")

    columns = [item.replace("_", " ").title() for item in columns]

    pdf.cell(w=30, h=10, txt=columns[0], border=1)
    pdf.cell(w=50, h=10, txt=columns[1], border=1)
    pdf.cell(w=35, h=10, txt=columns[2], border=1)
    pdf.cell(w=30, h=10, txt=columns[3], border=1)
    pdf.cell(w=30, h=10, txt=columns[4], ln=1, border=1)

    #total_price = 0

    for index, row in df.iterrows():
        pdf.set_font(family="Arial", size=10)
        pdf.set_text_color(80,80,80)
        pdf.cell(w=30, h=10, txt=str(row["product_id"]), border=1)
        pdf.cell(w=50, h=10, txt=str(row["product_name"]), border=1)
        pdf.cell(w=35, h=10, txt=str(row["amount_purchased"]), border=1)
        pdf.cell(w=30, h=10, txt=str(row["price_per_unit"]), border=1)
        pdf.cell(w=30, h=10, txt=str(row["total_price"]),ln=1, border=1)
        #total_price = total_price + row["total_price"]

    total_sum = df["total_price"].sum()
    pdf.cell(w=30, h=10, txt="", border=1)
    pdf.cell(w=50, h=10, txt="", border=1)
    pdf.cell(w=35, h=10, txt="", border=1)
    pdf.cell(w=30, h=10, txt="", border=1)
    pdf.cell(w=30, h=10, txt=str(total_sum),border=1, ln=2)

    pdf.cell(w=30, h=10, txt="", ln=1)

    pdf.set_font(family="Times", size=16, style="B")
    pdf.cell(w=30, h=12, txt=f"The total due amount is: {total_sum} AUD",ln=1)

    pdf.set_font(family="Times", size=16, style="B")
    pdf.cell(w=30, h=12, txt=f"PythonHow")
    pdf.image("pythonhow.png", w=10)

    pdf.output(f"PDFs/{filename}.pdf")

